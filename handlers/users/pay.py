from aiogram import types
from aiogram.types import Message
from data.config import ADMINS
from loader import dp, db, bot
from keyboards.default.korzinka_btn import location_btn
from keyboards.default.btns import menu_btn
from states.eltish import Yetkazish
from aiogram.dispatcher import FSMContext
from data.food_shablon import REGULAR_SHIPPING, FAST_SHIPPING, PICKUP_SHIPPING, foods_pay

from openpyxl import load_workbook

from docxtpl import DocxTemplate
from datetime import datetime as dt

hour = dt.now().strftime("%H")
minutes = dt.now().strftime("%M")
day = dt.now().strftime("%d")
month = dt.now().strftime("%m")
year = dt.now().strftime("%Y")

path = './evos_data/files/foods_data.xlsx'
wb_obj = load_workbook(path)
sheet_obj = wb_obj.active


@dp.message_handler(text="💰 To'lov qilish")
async def show_invoices(msg: Message):
    cl_class = await foods_pay(msg.from_user.id)
    await bot.send_invoice(chat_id=msg.from_user.id, **cl_class.generate_invoice(), payload="Kiyimlar")


@dp.shipping_query_handler()
async def choose_shipping(query: types.ShippingQuery):
    if query.shipping_address.country_code != "UZ":
        await bot.answer_shipping_query(shipping_query_id=query.id,
                                        ok=False,
                                        error_message="Chet elga yetkazib bera olmaymiz...")

    elif query.shipping_address.city.lower() == 'toshkent' or query.shipping_address.city.lower() == 'tashkent':
        await bot.answer_shipping_query(shipping_query_id=query.id,
                                        shipping_options=[
                                            FAST_SHIPPING, REGULAR_SHIPPING, PICKUP_SHIPPING],
                                        ok=True
                                        )
    else:
        await bot.answer_shipping_query(shipping_query_id=query.id,
                                        ok=False,
                                        error_message="Toshkentdan tashqariga yetkaza olmaymiz..."
                                        )


@dp.pre_checkout_query_handler()
async def process_pre_checkout_query(checkout_query: types.PreCheckoutQuery):
    await bot.answer_pre_checkout_query(pre_checkout_query_id=checkout_query.id, ok=True)
    user_id = checkout_query.from_user.id
    await bot.send_message(chat_id=user_id, text="Xaridingiz uchun rahmat")

    info_products = await db.show_korzinka(user_id) 
    info_text = ""
    
    for i in info_products:
        info_text += f"<b>{i['food_name'].title()}</b> <b>{i['food_count']} ta  {i['food_price']} so'mdan</b>  ID: {i['food_id']}\n"    

    
    
    desc = f"Quyidagi mahsulot sotildi:\n{info_text}\nID: {checkout_query.id}\n"
    desc += f"Xaridor: {checkout_query.order_info.name}, tel: {checkout_query.order_info.phone_number}\n"
    desc += f"Telegram User: <a href='https://t.me/{checkout_query.from_user.username}'>{checkout_query.from_user.full_name}</a>"
    await bot.send_message(chat_id=ADMINS[0], text=desc)
    
    
    
    data = DocxTemplate('evos_data/files/check.docx')
    
    obj = {
        'name': info_text
    }
    data.render(obj)
    data.save(f"evos_data/files/{checkout_query.from_user.full_name}-{day}-{month}-{year}.docx")
    
    ### Shu joyini o'zgartirish kerak
    with open(f"evos_data/files/{checkout_query.from_user.full_name}-{day}-{month}-{year}.docx", 'rb') as file:
        await bot.send_document(chat_id=user_id, document=file)
    
    
    id = sheet_obj.max_row + 1
    index = 0
    for i in range(id, len(info_products) + id):
        sheet_obj[f"A{i}"].value = user_id
        sheet_obj[f"B{i}"].value = info_products[index]["food_id"] 
        sheet_obj[f"C{i}"].value = info_products[index]["food_name"]
        sheet_obj[f"D{i}"].value = info_products[index]["food_price"]
        sheet_obj[f"E{i}"].value = info_products[index]["food_type"]
        sheet_obj[f"F{i}"].value = info_products[index]["food_count"]
        sheet_obj[f"F{i}"].value = checkout_query.from_user.username
        index += 1
    
    wb_obj.save("evos_data/files/foods_data.xlsx")
    
    for i in info_products:
        await db.delete_foods_from_korzinka(user_id)
    
    await bot.send_message(chat_id=user_id, text="Kuryerimiz aloqaga chiqishi uchun nomeringizni qoldiring")
    await Yetkazish.phone_number.set()

@dp.message_handler(state=Yetkazish.phone_number)
async def get_phone(msg: types.Message, state: FSMContext):
    await state.update_data(
        {
            "user_phone": msg.text
        }
    )
    await msg.answer("Locatsiyangizni tashlang", reply_markup=location_btn)
    await Yetkazish.location.set()

@dp.message_handler(state=Yetkazish.location, content_types=["location"])
async def get_phone(msg: types.Message, state: FSMContext):
    await state.update_data(
        {
            "latitude": msg.location.latitude,
            "longitude": msg.location.longitude
        }
    )
    await msg.answer("Turgan joyingiz haqida biror orintir yozib qoldiring", reply_markup=types.ReplyKeyboardRemove())
    await Yetkazish.orintir.set()

@dp.message_handler(state=Yetkazish.orintir)
async def get_phone(msg: types.Message, state: FSMContext):
    await state.update_data(
        {
            "orintir": msg.text
        }
    )
    await msg.answer("😊 Rahmat! Tez orada 🚚 kuryerimiz.\n🎁 Buyurtmangizni olib boradi", reply_markup=menu_btn)
    
    data = await state.get_data()
    phone_number = data.get("user_phone")
    latitude = data.get("latitude")
    longitude = data.get("longitude")
    orintir = data.get("orintir")
    
    await db.delete_korzinka(user_tg_id=msg.from_user.id)
    await db.add_to_customers(msg.from_user.id, msg.from_user.username, msg.from_user.full_name, phone_number, f"{latitude}:{longitude}", orintir)
    await state.finish()